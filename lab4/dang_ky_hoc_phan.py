# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

wb = load_workbook("D:\\third year\\python programing\\lab4\\excel.xlsx")

sheet = wb.active


def excel():
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 40
    sheet.column_dimensions["G"].width = 50

    sheet.cell(row=1, column=1).value = "MSSV"
    sheet.cell(row=1, column=1).value = "Họ tên"
    sheet.cell(row=1, column=2).value = "Ngày sinh"
    sheet.cell(row=1, column=3).value = "Email"
    sheet.cell(row=1, column=4).value = "Số điện thoại"
    sheet.cell(row=1, column=5).value = "Học kì"
    sheet.cell(row=1, column=6).value = "Năm học"
    sheet.cell(row=1, column=7).value = "Chọn môn học"


def focus1(event):
    course_field.focus_set()


def focus2(event):
    sem_field.focus_set()


def focus3(event):
    form_no_field.focus_set()


def focus4(event):
    contact_no_field.focus_set()


def focus5(event):
    email_id_field.focus_set()


def focus6(event):
    # set focus on the address_field box
    address_field.focus_set()


def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


def insert():
    if (
        name_field.get() == ""
        and course_field.get() == ""
        and sem_field.get() == ""
        and form_no_field.get() == ""
        and contact_no_field.get() == ""
        and email_id_field.get() == ""
        and address_field.get() == ""
    ):
        print("empty input")

    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        wb.save("D:\\third year\\python programing\\lab4\\excel.xlsx")

        name_field.focus_set()

        # call the clear() function
        clear()


if __name__ == "__main__":
    root = Tk()

    root.configure(background="light green")

    root.title("registration form")

    root.geometry("500x300")

    excel()

    heading = Label(root, text="Form", bg="light green")

    name = Label(root, text="Name", bg="light green")

    course = Label(root, text="Course", bg="light green")

    sem = Label(root, text="Semester", bg="light green")

    form_no = Label(root, text="Form No.", bg="light green")

    contact_no = Label(root, text="Contact No.", bg="light green")

    email_id = Label(root, text="Email id", bg="light green")

    address = Label(root, text="Address", bg="light green")

    heading.grid(
        row=0,
        column=1,
    )
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)
    var1 = IntVar()
    var2 = IntVar()
    var3 = IntVar()
    var4 = IntVar()

    # def print_selection():
    #     if (var1.get() == 1) & (var2.get() == 0):
    #         l.config(text="I love Python ")
    #     elif (var1.get() == 0) & (var2.get() == 1):
    #         l.config(text="I love C++")
    #     elif (var1.get() == 0) & (var2.get() == 0):
    #         l.config(text="I do not anything")
    #     else:
    #         l.config(text="I love both")

    c1 = Checkbutton(
        root,
        text="Lap trinh Python",
        variable=var1,
        onvalue=1,
        offvalue=0,
        # command=print_selection,
    )
    c1.grid(row=8, column=1)
    c2 = Checkbutton(
        root,
        text="Lap trinh Python",
        variable=var2,
        onvalue=1,
        offvalue=0,
        # command=print_selection,
    )
    c2.grid(row=8, column=2)
    c3 = Checkbutton(
        root,
        text="Lap trinh Python",
        variable=var3,
        onvalue=1,
        offvalue=0,
        # command=print_selection,
    )
    c3.grid(row=9, column=1)
    c4 = Checkbutton(
        root,
        text="Lap trinh Python",
        variable=var4,
        onvalue=1,
        offvalue=0,
        # command=print_selection,
    )
    c4.grid(row=9, column=2)

    name_field.bind("<Return>", focus1)

    course_field.bind("<Return>", focus2)

    sem_field.bind("<Return>", focus3)

    form_no_field.bind("<Return>", focus4)

    contact_no_field.bind("<Return>", focus5)

    email_id_field.bind("<Return>", focus6)

    name_field.grid(row=1, column=1, ipadx="100", columnspan=2)
    course_field.grid(row=2, column=1, ipadx="100", columnspan=2)
    sem_field.grid(row=3, column=1, ipadx="100", columnspan=2)
    form_no_field.grid(row=4, column=1, ipadx="100", columnspan=2)
    contact_no_field.grid(row=5, column=1, ipadx="100", columnspan=2)
    email_id_field.grid(row=6, column=1, ipadx="100", columnspan=2)
    address_field.grid(row=7, column=1, ipadx="100", columnspan=2)

    excel()

    submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert)
    submit.grid(row=10, column=1, columnspan=2, pady=10)

    root.mainloop()
